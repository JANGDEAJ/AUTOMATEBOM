"""
reporter.py - Excel + console reporting
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.drawing.image import Image as ExcelImage
import PIL.Image
from config import REPORT_DIR, REPORT_FILE, SCREENSHOT_DIR


# Colour fills
_GREEN  = PatternFill("solid", fgColor="C6EFCE")
_RED    = PatternFill("solid", fgColor="FFC7CE")
_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_HEADER = PatternFill("solid", fgColor="4472C4")
_THIN   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def save_report(results: list[dict]):
    """Save results list to Excel with colour-coded Status column in all project directories."""
    os.makedirs(REPORT_DIR, exist_ok=True)

    # SAFETY DUPLICATE BACKUP: Always backup existing Excel file before modifying
    if os.path.exists(REPORT_FILE):
        try:
            import shutil
            bak_path = REPORT_FILE + f".bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            shutil.copy(REPORT_FILE, bak_path)
            shutil.copy(REPORT_FILE, REPORT_FILE + ".safety_backup.xlsx")
        except Exception as e:
            print(f"[BACKUP WARNING] Could not create safety backup: {e}")

    df = pd.DataFrame(results)

    # Reorder columns neatly
    ordered = ["TC ID", "Module", "Function", "Role", "Permission Type",
                "Test Type", "ขั้นตอนทดสอบ", "ผลที่คาดหวัง", "Status", "Comments",
                "Screenshot"]
    for col in ordered:
        if col not in df.columns:
            df[col] = ""
    extra = [c for c in df.columns if c not in ordered]
    df = df[ordered + extra]

    df.to_excel(REPORT_FILE, index=False, sheet_name="Results")

    # Style the workbook
    wb = load_workbook(REPORT_FILE)
    ws = wb["Results"]

    # Header row
    for cell in ws[1]:
        cell.fill   = _HEADER
        cell.font   = Font(bold=True, color="FFFFFF")
        cell.border = _THIN
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows and images
    status_col = df.columns.get_loc("Status") + 1
    screenshot_col = df.columns.get_loc("Screenshot") + 1 if "Screenshot" in df.columns else None

    if screenshot_col:
        ws.column_dimensions[get_column_letter(screenshot_col)].width = 25

    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        status_val = row[status_col - 1].value or ""
        fill = _GREEN if status_val == "Passed" else (_RED if status_val == "Failed" else _YELLOW)
        for cell in row:
            cell.border    = _THIN
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        row[status_col - 1].fill = fill

        # Handle screenshot embedding
        if screenshot_col:
            ws.row_dimensions[idx].height = 100  # Adjust row height for image
            ss_val = row[screenshot_col - 1].value
            if ss_val:
                ss_path = str(ss_val)
                if not os.path.isabs(ss_path):
                    ss_path = os.path.join(SCREENSHOT_DIR, ss_path)
                
                if os.path.exists(ss_path):
                    try:
                        # Keep text for pandas, but make it invisible
                        row[screenshot_col - 1].font = Font(color="FFFFFF")
                        # Add image
                        img = ExcelImage(ss_path)
                        # Scale down image to fit cell (approx 130px height)
                        img.height = 125
                        img.width = 125 * (img.width / img.height) # maintain aspect ratio
                        
                        # Anchor image to the cell
                        cell_ref = f"{get_column_letter(screenshot_col)}{idx}"
                        ws.add_image(img, cell_ref)
                    except Exception as e:
                        row[screenshot_col - 1].value = f"Img Error: {str(e)}"
                else:
                    row[screenshot_col - 1].value = "Image not found"

    _autofit(ws)
    if screenshot_col:
        ws.column_dimensions[get_column_letter(screenshot_col)].width = 25 # Ensure width again after autofit
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    total   = len(df)
    passed  = (df["Status"] == "Passed").sum()
    failed  = (df["Status"] == "Failed").sum()
    skipped = (df["Status"] == "Skipped").sum()

    ws_sum.append(["BOM UAT Automation Report"])
    ws_sum.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws_sum.append([])
    ws_sum.append(["", "Count", "Percent"])
    ws_sum.append(["Total",   total,   "100%"])
    ws_sum.append(["Passed",  passed,  f"{passed/total*100:.1f}%" if total else "-"])
    ws_sum.append(["Failed",  failed,  f"{failed/total*100:.1f}%" if total else "-"])
    ws_sum.append(["Skipped", skipped, f"{skipped/total*100:.1f}%" if total else "-"])

    # Per-permission breakdown
    ws_sum.append([])
    ws_sum.append(["Per Permission Type", "Total", "Passed", "Failed"])
    for ptype, grp in df.groupby("Permission Type"):
        p = (grp["Status"] == "Passed").sum()
        f = (grp["Status"] == "Failed").sum()
        ws_sum.append([ptype, len(grp), p, f])

    # Per-role breakdown
    ws_sum.append([])
    ws_sum.append(["Per Role", "Total", "Passed", "Failed"])
    for role, grp in df.groupby("Role"):
        p = (grp["Status"] == "Passed").sum()
        f = (grp["Status"] == "Failed").sum()
        ws_sum.append([role, len(grp), p, f])

    ws_sum["A1"].font = Font(bold=True, size=14)
    _autofit(ws_sum)

    wb.save(REPORT_FILE)

    print(f"\n Report saved -> {REPORT_FILE}")
    print(f"   Total: {total}  Passed: {passed}  Failed: {failed}  Skipped: {skipped}")
    return REPORT_FILE
